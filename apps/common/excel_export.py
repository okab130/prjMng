from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


class ExcelExporter:
    """Excel出力共通クラス"""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # スタイル定義
        self.header_font = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def set_header(self, headers):
        """ヘッダー行を設定"""
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def add_row(self, row_num, data):
        """データ行を追加"""
        for col_num, value in enumerate(data, 1):
            cell = self.ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = self.border
            cell.alignment = Alignment(vertical='center')
    
    def auto_adjust_column_width(self):
        """列幅を自動調整"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def get_response(self, filename):
        """HTTPレスポンスを生成"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        self.wb.save(response)
        return response
